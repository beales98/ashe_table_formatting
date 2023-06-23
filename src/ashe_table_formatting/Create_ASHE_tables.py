# -*- coding: utf-8 -*-
"""
@author: Paul Dunstan
@Customer: ASHE Team
"""
import pandas as pd
import os
import numpy as np
import openpyxl as opy
from openpyxl.utils.dataframe import dataframe_to_rows
import copy
import collections.abc as c
from openpyxl.styles import PatternFill
from decimal import Decimal, ROUND_HALF_UP

from ashe_table_formatting.pipeline_config import *

def all_keys(dict_obj):
    ''' This function generates all keys of
        a nested dictionary. 
        
        Parameters
        ----------
        dict_obj: Dictionary
            Dictionary to retreive key from.
            
    '''
    # Iterate over all keys of the dictionary
    for key , value in dict_obj.items():
        yield key
        # If value is of dictionary type then yield all keys
        # in that nested dictionary
        if isinstance(value, dict):
            for k in all_keys(value):
                yield k

def deep_update(source, overrides):
    '''
    Update a nested dictionary or similar mapping.
    Modify ``source`` in place.
    
    Parameters
    ----------
    source: Dictionary
        Source dictionary.
    overrides : Dictionary
        New element for the dictionary.

    Returns
    -------
    source : Combined dictionary. 
    '''
    for key, value in overrides.items():
        
        if isinstance(value, c.Mapping) and value:
            returned = deep_update(source.get(key, {}), value)
            source[key] = returned
            
        else:
            source[key] = overrides[key]
            
    return source

def get_files_from_list(csv_path, file_list):
    '''
    This function a locates files from get files to a nested dictionary

    Parameters
    ----------
    csv_path : String
        Path to the input CSV files
    file_list : List of strings
        List of input file names

    Returns
    -------
    result : Nested Dictionary
        Nested dictionary containing all datasets used to create sub-table.

    '''
    os.chdir(csv_path)
    result = {}

    for file_name in file_list:
        datasets = find_datasets(csv_path, file_name)
        
    
        for csv in datasets:
            short = csv.replace('.csv', '') 
            split = short.split(' - ')
            data_file_dict = {split[0] : {split[1] : {split[2] : get_files(csv, file_name)}}}
            deep_update(result, data_file_dict)
    
    return result
             
def get_files(csv, file_name):
    '''
    This function retrieves the reads csv files and manipulates them into dataframes.

    Parameters
    ----------
    csv : String
        csv file path to be ingested.
    file_name : String
        name of csv file - shorthand.

    Returns
    -------
    data : List of dataframes
        9 Sorted dataframes with data from the 9 employee types.

    '''
    data = pd.read_csv(csv, header = 5, encoding = 'unicode_escape', thousands=',', on_bad_lines = 'warn', skip_blank_lines = False)
    end_section_row = data[data.iloc[:, 0].str.contains("key1=") == True].index + 6
    end_section_row = end_section_row.insert(0,0)
    end_section_row = end_section_row.insert(9,data.index[-1])
    
    data = [data.iloc[end_section_row[i]:end_section_row[i+1],:] for i in range(0,9)]
    data = [data[i].dropna() for i in range(0,9)]
    data = [data[i].reset_index(drop=True) for i in range(0,9)]
    [data[i].rename(columns = {data[i].columns[0]: 'Description'}, inplace = True) for i in range(0,9)]
    [data[i].rename(columns = {data[i].columns[1]: 'Code'}, inplace = True) for i in range(0,9)]
    [data[i].replace(',','', regex=True, inplace=True) for i in range(0,9)]
    
    for i in range(0,9):
        data[i]['Code'] = data[i]['Code'].astype('str').replace(r'\.\d+$', '', regex=True)
        data[i]['Code'] = data[i]['Code'].apply(str)
        data[i]['Code'] = f'{file_name}' + ' ' + data[i]['Code'].astype(str)
    
    
    key_list = [*Employee_key]
    
    data = {key_list[x]: data[x] for x in range(0,9)}

    return data

def find_datasets(csv_path, file_name):
    '''
    Finds all csv files in path and appends them to a list if they are
    required by file_name.

    Parameters
    ----------
    csv_path : String
        File path to CSVs.
    file_name : String
        file names to be loaded.

    Returns
    -------
    csv_list : List of String
        List of the CSV files needed.

    '''
    csv_list =[]
    
    for file in os.listdir(csv_path):
        if file_name in file:
            csv_list.append(file)
            
    return csv_list

def get_order_of_variables(template_xlsx, table_file_names):
    '''
    Pulls out the order of varaibles from the templates. ORgansises input data
    into a nested dictionary so that it can be called simply.

    Parameters
    ----------
    template_xlsx : Workbook
        Loaded template workbook for relevant table.
    table_file_names : List of String
        List of all datasets that a table requries. Also as tabs on template.

    Returns
    -------
    ordered_df : Dataframe
        Dataframe of variable names and their order ready to merge with input
        data.
    '''
    order = {}
    for table in table_file_names:
        sheet = template_xlsx[table]
        col_key = sheet['A']
        col_order = sheet['C']
        list_key = []
        list_order = []
        
        for col in col_key:
            list_key.append(col.value)
        
        for col in col_order:
            list_order.append(col.value)
        
        list_key = [f'{table}' + ' ' + str(x) for x in list_key]
        
        order = deep_update(order, dict(zip(list_key[1:], list_order[1:])))
        
        
    ordered_list = list(order.items())
    ordered_df = pd.DataFrame(ordered_list).rename(columns = {0: 'Code', 1: 'Order'})
            
    return ordered_df

def create_data_ready(csv_path, table_name, variable, type_of_value, employee_type, year):
    '''
    This function loads in the required datasets for a table and processes them
    ready to be used.
    
    We must convert missing data to nans first before we can work with it.

    Parameters
    ----------
    csv_path : String
        Path to input data.
    table_name : String
        Name of table to get data for.
    variable : String
        Name of variable to get data for.
    type_of_value : String
        Type of value to get data for.
    employee_type : String
        Type of employee to get data for.
    year : Numeric
        Current Year.

    Returns
    -------
    data_needed : Dataframe
        A single dataframe with the relevant data.

    '''
    
    table_file_names = Published_tables_data[table_name] # Load table file names
    table_data = get_files_from_list(csv_path, table_file_names) #get the input data ready
    variable_needed = Published_table_breakdown[variable] # get variable for loop
    numeric_cols = ['population number', '10', '20', '25', '30', '40', '60', '70', '75', '80', '90']
    percentiles = ['10', '20', '25', '30', '40', '60', '70', '75', '80', '90']
    combined_data = []
    
    '''Get the data from the SAS input sheets'''
    for file in table_file_names:

        combined_data.append(table_data[file][variable_needed + ' ' + f'{year}'][type_of_value][employee_type])
        data_needed = pd.concat(combined_data)
        data_needed = data_needed.drop_duplicates()
    
    '''Process the input data'''
    data_needed['population number'] = data_needed['population number'].replace(r'\.(?=\s|$)','0', regex=True)
    data_needed[numeric_cols] = data_needed[numeric_cols].apply(lambda x : pd.to_numeric(x, errors='coerce'))
    
    for i in percentiles:
        data_needed[i] = data_needed[i].map(lambda x: Decimal(str(x)).quantize(Decimal('.1'), rounding=ROUND_HALF_UP)) # Round our percentiles to 1dp
    
    if type_of_value == 'Values':
        data_needed['population number'] = data_needed['population number'] / 1000
        data_needed['population number'] = data_needed['population number'].map(lambda x: Decimal(str(x)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    
    data_needed['Mean'] = data_needed['Mean'].replace(r'\.(?=\s|$)', np.nan, regex=True)
    data_needed['Mean'] = pd.to_numeric(data_needed['Mean']).map(lambda x: Decimal(str(x)).quantize(Decimal('.1'), rounding=ROUND_HALF_UP))
    data_needed['Mean'] = data_needed['Mean'].replace(np.nan, '.', regex=True)
    
    data_needed['Median'] = data_needed['Median'].replace(r'\.(?=\s|$)', np.nan, regex=True)
    
    # Convert pence to pounds for hourly datasets    
    if variable == 'Hourly Pay' or variable == 'Hourly pay - Excluding overtime':
        if type_of_value == 'Values':
            print('Converting pence to pound')
            cols_pence = ['Median', '10', '20', '25', '30', '40', '60', '70', '75', '80', '90']
            for i in cols_pence:
                try:
                    data_needed[i] = pd.to_numeric(data_needed[i].map(lambda x: Decimal(str(x/100)).quantize(Decimal('.10'), rounding=ROUND_HALF_UP)))
                except(TypeError):
                    data_needed[i] = pd.to_numeric(data_needed[i], errors='ignore')
                    data_needed[i] = data_needed[i].map(lambda x: Decimal(str(x/100)).quantize(Decimal('.10'), rounding=ROUND_HALF_UP))
    else:           
        data_needed['Median'] = pd.to_numeric(data_needed['Median']).map(lambda x: Decimal(str(x)).quantize(Decimal('.1'), rounding=ROUND_HALF_UP))
    
    data_needed['Median'] = data_needed['Median'].replace(np.nan, '.', regex=True) # Replace missing data with '.' for median col
    
    for col in numeric_cols:
        data_needed[col] = pd.to_numeric(data_needed[col], errors = 'coerce')
        data_needed[col] = data_needed[col].replace(np.nan, '.', regex=True) # Replace missing data with '.'

    return data_needed

def copy_sheet_style(ws_template: opy.Workbook.worksheets, ws_published: opy.Workbook.worksheets, isvalmain):
    '''
    Function copies cell styles from a template worksheet to the output
    worksheet. Has a seperate function for the value main sheet that requires
    x where the CV is greater than 20.
    
    Parameters
    ----------
    ws_template -- Loaded worksheet with desired cell styles.
    ws_published -- Worksheet to receive the cell style.
    '''
    if isvalmain == 'valmain':
        for row in (ws_template.rows):
            for cell in row:
                new_cell = ws_published.cell(row=cell.row, column=cell.col_idx)
                
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)
                    
                if cell.col_idx > 2 and cell.value == 'x':
                    new_cell.value = 'x'
    else:
        for row in (ws_template.rows):
            for cell in row:
                new_cell = ws_published.cell(row=cell.row, column=cell.col_idx)
                
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)   
            
def get_workbook_name(template_name, table_name, table_variable, value_type, year):
    '''
    This function is used to name the different workbooks for saving to excel.

    Parameters
    ----------
    template_name : String
        Template name of table being procesed.
    table_name : String
        Table being processed.
    table_variable : String
        Type of value to find name for.
    value_type : String
        Type of value to find name for.
    year : Numeric
        Current year.

    Returns
    -------
    complete_name : String
        Name to save the workbook as.
    '''
    table_sub_number = Table_sub_number_key[table_variable]
    name = table_name.split(' - ')[0]
    
    if value_type == 'Values':
        complete_name = name + table_sub_number + 'a' + ' ' + table_variable
    else:
        complete_name = name + table_sub_number + 'b' + ' ' + table_variable
    
    return complete_name
  
def compare_year_and_previous_year(data, data_py):
    '''
    Function that calucated the percentage change between the Mean and median
    of two years

    Parameters
    ----------
    data : Dataframe
        Dataframe of this years table data - Main.
    data_py : Dataframe
        Dataframe of last years table data - Main.

    Returns
    -------
    data : Dataframe
        Completed dataframe with the comparisons loaded.

    '''
    data['Median'] = pd.to_numeric(data['Median'], errors='coerce')
    data['Mean'] = pd.to_numeric(data['Mean'], errors='coerce')
    data_py['Median'] = pd.to_numeric(data_py['Median'], errors='coerce')
    data_py['Mean'] = pd.to_numeric(data_py['Mean'], errors='coerce')
    
    if data['Code'].equals(data_py['Code']) == True:
        data['Year on Year % Change'] = percentage_change(data_py['Median'], data['Median'])
        data['Year on Year % Change.1'] = percentage_change(data_py['Mean'], data['Mean'])
        
        data['Year on Year % Change'] = data['Year on Year % Change'].replace(np.nan, '')
        data['Year on Year % Change.1'] = data['Year on Year % Change.1'].replace(np.nan, '')
        
    return data

def percentage_change(col1,col2):
    '''
    Calulates the percentage change between two columns.

    Parameters
    ----------
    col1 : Dataframe column
        Column of a dataframe.
    col2 : Dataframe column
        Column of a dataframe.

    Returns
    -------
    df_col: Dataframe column
        Column of a dataframe.

    '''
    df_col = (((col2 - col1) / col1) * 100)
    
    try:
        df_col = df_col.map(lambda x: Decimal(str(x)).quantize(Decimal('.1'), rounding=ROUND_HALF_UP))
    except:
        df_col = df_col.replace(np.inf, np.nan)
        df_col = df_col.map(lambda x: Decimal(str(x)).quantize(Decimal('.1'), rounding=ROUND_HALF_UP))

    return df_col

def apply_order(data, template_order):
    '''
    Merges the prepared input data to the order data from the template. Any
    missing data that is in the template but not in the SAS input data is
    labeled as :

    Parameters
    ----------
    data : Dataframe
        Datatrame of the prepared input data.
    template_order : Dataframe
        Dataframe of the corresponding codes and order for input data.

    Returns
    -------
    data : Dataframe
        Merged dataframe with input data and order.

    '''
    cols_to_empty = ['Median','Year on Year % Change', 'Mean', 'Year on Year % Change.1', '10', '20', '25', '30', '40', '60', '70', '75', '80', '90']
    data = pd.merge(template_order, data, how= 'outer', on = 'Code')
    data = data[data['Order'].notna()]
    data = data.sort_values('Order')
    data = data.fillna(':')
    data[cols_to_empty] = data[cols_to_empty].replace(':', '')
    return data

def make_safe(data):
    '''
    Remove unsafe data frame a dataframe looking for 0 in the safe column

    Parameters
    ----------
    data : Dataframe
        Datafeame to make safe.

    Returns
    -------
    data : Dataframe
        Safe Dataframe.

    '''
    cols_to_empty = ['Median','Year on Year % Change', 'Mean', 'Year on Year % Change.1', '10', '20', '25', '30', '40', '60', '70', '75', '80', '90']
    data = data.mask(data['Safe'] == 0, np.nan)
    data = data.fillna('..')
    data[cols_to_empty] = data[cols_to_empty].replace('..', '')
    return data

def drop_cols(data):
    '''
    Drops columns that are not published

    Parameters
    ----------
    data : Dataframe
        Dataframe to be published but has excess columns.

    Returns
    -------
    data : Dataframe
        Dataframe to be published with cols removed.

    '''
    data.drop('Code', inplace = True, axis = 1)
    data.drop('Order', inplace = True, axis = 1)
    data.drop('Description', inplace = True, axis = 1)
    data.drop('Safe', inplace = True, axis = 1)
    return data

def create_workbook(csv_path, csv_previous_year_path, template_path, output_path, table_name, table_variable, year):
    '''
    Function creates a workbooks for a subtable. This will have 5 excel
    outputs unsafe, safe and main for Values and Unsafe and main for CV.
    Applies most of the function above to do this

    Parameters
    ----------
    csv_path : String
        Path to CSV.
    csv_previous_year_path : String
        Path to previous years CSV.
    template_path : String
        Path to templates.
    output_path : String
        Path to where to store outputs.
    table_name : String
        Table to process.
    table_variable : String
       Variable to process.
    year : numeric
        Current Year.

    Returns
    -------
    5 Workbooks.

    '''
    
    '''Create sub directory in output path for table name'''
    if not os.path.exists(output_path + '/' + table_name):
        os.mkdir(output_path + '/' + table_name)
    output_path = output_path + '/' + table_name
    
    '''Load up templates ready to be filled with data'''
    template_cv_main = opy.load_workbook(filename = template_path + '/' + Published_tables_templates[table_name]) #This version will have x where cv >20 to transfer to value sheet
    template_val_main = opy.load_workbook(filename = template_path + '/' + Published_tables_templates[table_name])
    template_cv_unsafe = opy.load_workbook(filename = template_path + '/' + Published_tables_templates[table_name])
    template_val_unsafe = opy.load_workbook(filename = template_path + '/' + Published_tables_templates[table_name])
    template_val_safe = opy.load_workbook(filename = template_path + '/' + Published_tables_templates[table_name])
    template_cv_main_final = opy.load_workbook(filename = template_path + '/' + Published_tables_templates[table_name]) # A special final version that will not have x where cv >20
    
    '''Load up the correct footnote'''
    template_footnote = pd.read_excel(template_path + '/' + 'Footnotes template.xlsx', sheet_name = ['FootNotes1', 'FootNotes2', 'FootNotes3', 'FootNotes4'])
    if table_variable == 'Annual pay - Gross':
        footnote = template_footnote['FootNotes2']
    if table_variable == 'Overtime pay':
        footnote = template_footnote['FootNotes3']
    if table_variable == 'Annual pay - Incentive':
        footnote = template_footnote['FootNotes4']
    else:
        footnote = template_footnote['FootNotes1']
    
    '''Gather variables for the loops and for naming the workbooks'''
    template_sheet_names = list(Employee_key.values()) # The tabs on the sheet that will be filled with data.
    employee_list = [*Employee_key] # A list of the employee types
    template_name = Published_tables_templates[table_name].split(' template')[0] # Getting the name of the table from the template
    workbook_name_cv = get_workbook_name(template_name, table_name, table_variable, 'CVs', year) # Get CV workbook name for saving
    workbook_name_val = get_workbook_name(template_name, table_name, table_variable, 'Values', year) # Get Val workbook name for saving
    drop_list = Published_tables_data[table_name] # The tabs on the template that need to be dropped for publication
    columns_to_mask = ['population number','Year on Year % Change', 'Year on Year % Change.1', '10', '20', '25', '30', '40', '60', '70', '75', '80', '90'] # Columns that need to be x when val population <3
    
    '''Get order from template'''
    template_order = get_order_of_variables(template_cv_main, Published_tables_data[table_name]) # Get the codes and order from the template
    max_order = template_order['Order'].max() # Find out which is the max order (for working out where the footnotes go)
    
    '''Begin creating sub tables'''
    print('Doing ' + table_name + ' ' + table_variable)
    print('Loading data')
    
    data_list_cv = []
    data_list_val = []
    data_list_val_py = []
    
    '''Load data and round as required'''
    for i in range (0,9):
        print('Loading ' + table_variable + ' ' + f'{employee_list[i]}' + ' CVs')
        data_list_cv.append(create_data_ready(csv_path, table_name, table_variable, 'CVs', employee_list[i], year))
        
        print('     Loading ' + table_variable + ' ' + f'{employee_list[i]}' + ' Values')
        data_list_val.append(create_data_ready(csv_path, table_name, table_variable, 'Values', employee_list[i], year))
        
        print('         Loading previous year ' + table_variable + ' ' + f'{employee_list[i]}')
        data_list_val_py.append(create_data_ready(csv_previous_year_path, table_name, table_variable, 'Values', employee_list[i], year-1))
        
        data_val_main_masked = data_list_val.copy()
        
    print('Data_loaded')
    
    '''Loop through employee type to process and write data'''
    for i, sheet in enumerate(template_sheet_names):
        
        '''Apply Ordering from template'''
        print('Writing ' + f'{year}' + ' ' + sheet + ' data')
        
        data_cv =  apply_order(data_list_cv[i], template_order)
        data_val =  apply_order(data_list_val[i], template_order)
        data_val_py =  apply_order(data_list_val_py[i], template_order)
        
        '''A Special version is needed for values main where x is applied across rows with population <=3 '''
        data_val_main_masked[i][columns_to_mask] = data_val_main_masked[i][columns_to_mask].mask(data_val_main_masked[i]['population number'] <= 3, 'x')
        data_val_main = apply_order(data_val_main_masked[i], template_order)        
        
        '''Make Copies'''
        data_cv_unsafe = data_cv.copy()
        data_val_unsafe = data_val.copy()
        data_val_safe = data_val.copy()
        data_cv_main = data_cv.copy()
        data_val_main = data_val_main.copy()
        
        '''Add comparisons'''
        data_val_main = compare_year_and_previous_year(data_val_main, data_val_py)
        data_val_main = data_val_main.sort_values('Order')
        
        '''Do safe'''
        data_val_safe = make_safe(data_val_safe)   
        
        data_val_safe = drop_cols(data_val_safe)
        
        rows_val_safe = dataframe_to_rows(data_val_safe, index=False, header=False)
        
        '''Do unsafe'''    
        data_cv_unsafe = drop_cols(data_cv_unsafe)
        data_val_unsafe = drop_cols(data_val_unsafe)
        
        rows_cv_unsafe = dataframe_to_rows(data_cv_unsafe, index=False, header=False)
        rows_val_unsafe = dataframe_to_rows(data_val_unsafe, index=False, header=False)
        
        '''Do main'''
        
        data_cv_main = make_safe(data_cv_main)
        data_val_main = make_safe(data_val_main)
        
        data_cv_main = drop_cols(data_cv_main)
        data_val_main = drop_cols(data_val_main)
        
        data_cv_main = data_cv_main.apply(lambda x: pd.to_numeric(x, errors='ignore'))
        data_val_main = data_val_main.apply(lambda x: pd.to_numeric(x, errors='ignore'))
        
        rows_cv_main = dataframe_to_rows(data_cv_main, index=False, header=False)
        rows_val_main = dataframe_to_rows(data_val_main, index=False, header=False)
        
        '''Create a spare CV for the final version (without x's)'''
        rows_cv_main_final = dataframe_to_rows(data_cv_main, index=False, header=False)
        
        '''Print unsafe to excel'''
        sheet_active_cv_unsafe = template_cv_unsafe[sheet]
        sheet_active_cv_unsafe['A1'] = workbook_name_cv
        
        sheet_active_val_unsafe = template_val_unsafe[sheet]
        sheet_active_val_unsafe['A1'] = workbook_name_val
        
        for r_idx, row in enumerate(rows_cv_unsafe, 1):
            for c_idx, value in enumerate(row, 1):
                sheet_active_cv_unsafe.cell(row=r_idx + 5, column=c_idx + 2, value=value) # print data for CV
                
        for i in range(0,6):
            sheet_active_cv_unsafe.cell(row= max_order + 5 + 1 + i , column= 1, value = footnote['Footnote'].iat[i]) # add footnote     
                
        for r_idx, row in enumerate(rows_val_unsafe, 1):
            for c_idx, value in enumerate(row, 1):
                sheet_active_val_unsafe.cell(row=r_idx + 5, column=c_idx + 2, value=value) # print data for Value
                
        for i in range(0,6):
            sheet_active_val_unsafe.cell(row=max_order + 5 + 1 + i , column= 1, value = footnote['Footnote'].iat[i]) # add footnote 
        
        copy_sheet_style(template_cv_main[sheet], sheet_active_cv_unsafe, 'notvalmain') # Copy template formatting to CV
        copy_sheet_style(template_cv_main[sheet], sheet_active_val_unsafe, 'notvalmain') # Copy template formatting to Value
       
        print('Done ' + ' ' + sheet + ' Unsafe')
        
        '''Print safe to excel'''
        sheet_active_val_safe = template_val_safe[sheet]
        sheet_active_val_safe['A1'] = workbook_name_val
        
        for r_idx, row in enumerate(rows_val_safe, 1):
            for c_idx, value in enumerate(row, 1):
                sheet_active_val_safe.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                
        for i in range(0,6):
            sheet_active_val_safe.cell(row=max_order + 5 + 1 + i , column= 1, value = footnote['Footnote'].iat[i])
        
        copy_sheet_style(template_cv_main[sheet], sheet_active_val_safe, 'notvalmain')
        print('Done ' + ' ' + sheet + ' Safe')
        
        '''Print main to excel'''
        sheet_active_cv_main = template_cv_main[sheet]
        sheet_active_cv_main['A1'] = workbook_name_cv
        
        sheet_active_val_main = template_val_main[sheet]
        sheet_active_val_main['A1'] = workbook_name_val

        print('Done ' + ' ' + sheet + ' Main')
        
        '''Printing CVs main and formatting'''
        light_blue = PatternFill(patternType = 'gray0625', start_color = '00FFFF', end_color = '00FFFF', fill_type= 'solid')
        dark_blue = PatternFill(start_color = '33CCCC', end_color = '00FFFF', fill_type= 'solid')
        
        for r_idx, row in enumerate(rows_cv_main, 1):
            for c_idx, value in enumerate(row, 1):
                try:
                    if sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).value == 0: 
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
                    elif 0 < sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 5:
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value)   
                        
                    elif 5 < sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 10:
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).fill = light_blue
                        
                    elif 10 < sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 20:
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).fill = dark_blue
                        
                    elif sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).value > 20:
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value= 'x').fill = dark_blue
                        
                    else:
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
                except:
                    if sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value).value == '.' and c_idx >= 6:
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value= 'x').fill = dark_blue
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value= 'x')
                        
                    else:
                        sheet_active_cv_main.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
        for i in range(0,6):
            sheet_active_cv_main.cell(row=max_order + 5 + 1 + i , column= 1, value = footnote['Footnote'].iat[i])
            
        '''Do a sheet for CV without x's'''
        
        sheet_active_cv_main_final = template_cv_main_final[sheet]
        sheet_active_cv_main_final['A1'] = workbook_name_cv
        
        for r_idx, row in enumerate(rows_cv_main_final, 1):
            for c_idx, value in enumerate(row, 1):
                try:
                    if sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value == 0: 
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
                    elif 0 < sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 5:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)   
                        
                    elif 5 < sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 10:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).fill = light_blue
                        
                    elif 10 < sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 20:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).fill = dark_blue
                        
                    elif sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value > 20:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value= value).fill = dark_blue
                        
                    else:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
                except:
                    if sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value == '.' and c_idx >= 6:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value= 'x').fill = dark_blue
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value= 'x')
                    
                    else:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
        for i in range(0,6):
            sheet_active_cv_main_final.cell(row=max_order + 5 + 1 + i , column= 1, value = footnote['Footnote'].iat[i])
        
        '''Printing Values main and formatting'''
        for r_idx, row in enumerate(rows_val_main, 1):
           for c_idx, value in enumerate(row, 1):
                sheet_active_val_main.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                   
        for i in range(0,6):
            sheet_active_val_main.cell(row=max_order + 5 + 1 + i , column= 1, value = footnote['Footnote'].iat[i]) 
               
    '''Drop unneeded tabs'''
    for sheet in drop_list:
        del template_cv_unsafe[sheet]
        del template_val_unsafe[sheet]
        del template_val_safe[sheet]
        del template_cv_main[sheet]
        del template_val_main[sheet]
        del template_cv_main_final[sheet]
        
    '''Save all our workbooks'''
    template_cv_unsafe.save(output_path + '/' + template_name + ' ' + workbook_name_cv + ' ' + f'{year}' + ' ' +  'CV' + ' ' + 'Unsafe' + '.xlsx')

    template_val_unsafe.save(output_path + '/' + template_name + ' ' + workbook_name_val + ' ' + f'{year}' + ' ' + 'Unsafe' + '.xlsx')
    
    template_val_safe.save(output_path + '/' + template_name + ' ' + workbook_name_val + ' ' + f'{year}' + ' ' + 'Safe' + '.xlsx')
    
    template_cv_main.save(output_path + '/' + template_name + ' ' + workbook_name_cv + ' ' + f'{year}' +  ' ' +  'CV'  + '.xlsx')
    
    '''Copy CV formatting to Values'''
    for sheet in template_sheet_names:
        copy_sheet_style(template_cv_main[sheet], template_val_main[sheet], 'valmain')
        
    template_val_main.save(output_path + '/' + template_name + ' ' + workbook_name_val + ' ' + f'{year}' + '.xlsx')
    
    '''Redo CV now to remove X for over 20 scores'''
    
    for i, sheet in enumerate(template_sheet_names):
        
        sheet_active_cv_main_final = template_cv_main_final[sheet]
        sheet_active_cv_main_final['A1'] = workbook_name_cv
        
        for r_idx, row in enumerate(rows_cv_main_final, 1):
            for c_idx, value in enumerate(row, 1):
                try:
                    if sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value == 0: 
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
                    elif 0 < sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 5:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)   
                        
                    elif 5 < sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 10:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).fill = light_blue
                        
                    elif 10 < sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value <= 20:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).fill = dark_blue
                        
                    elif sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value > 20:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value= value).fill = dark_blue
                        
                    else:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)
                        
                except:
                    if sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value).value == '.' and c_idx >= 6:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value= 'x').fill = dark_blue
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value= 'x')
                    
                    else:
                        sheet_active_cv_main_final.cell(row=r_idx + 5, column=c_idx + 2, value=value)
        
        for i in range(0,6):
            sheet_active_cv_main_final.cell(row=max_order + 5 + 1 + i , column= 1, value = footnote['Footnote'].iat[i]) 
    
    template_cv_main_final.save(output_path + '/' + template_name + ' ' + workbook_name_cv + ' ' + f'{year}' +  ' ' +  'CV'  + '.xlsx')
    
    print('Saved all workbooks')
        
def create_table(csv_path, csv_previous_year_path, template_path, output_path, table_name, year):
    '''
    This function loops though the sub.tables to create all 11 of them.

    Parameters
    ----------
    csv_path : String
        Path to CSV.
    csv_previous_year_path : String
        Path to previous year csv.
    template_path : String
        Path to templates.
    output_path : String
        Path to where outputs will be saved.
    table_name : String
        Table to process.
    year : numeric
        Current year.

    Returns
    -------
    Saved excel documents for a table

    '''
    table_file_names = Published_tables_data[table_name]
    print('Table contains - ')
    print(table_file_names)
    
    table_variable_list = list(Published_table_breakdown.keys())
    
    for table_variable in table_variable_list:
        create_workbook(csv_path, csv_previous_year_path, template_path, output_path, table_name, table_variable, year)
        
# Testing not used in main code
